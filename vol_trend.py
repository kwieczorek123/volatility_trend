import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression
import numpy as np

# Define file names for input data
file_names = ['vol_trend_data/EURUSD_chop.csv', 'vol_trend_data/GBPUSD_chop.csv', 'vol_trend_data/USDJPY_chop.csv',
              'vol_trend_data/XAUUSD_chop.csv']
spread_all_symbols_file = 'vol_trend_data/spread_all_symbol.csv'
execution_spread_file = 'vol_trend_data/execution_spread.csv'
pnl_files = ['vol_trend_data/EURUSD_pnl.csv', 'vol_trend_data/GBPUSD_pnl.csv', 'vol_trend_data/USDJPY_pnl.csv',
             'vol_trend_data/XAUUSD_pnl.csv']

# Create a dictionary to store PnL data for each symbol
pnl_df_dict = {symbol: pd.read_csv(pnl_file, parse_dates=['day']) for pnl_file, symbol in
               zip(pnl_files, ['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD'])}

# Define date range and rolling window length
start_date = '2021-01-01'
end_date = '2023-02-28'
roll = 252


# Function to add volatility columns to the input DataFrame
def add_volatility_columns(df):
    df['Volatility'] = (df['high'] - df['low']) / df['low']
    df['Volatility_median'] = df['Volatility'].rolling(window=roll).median()
    df['Volatility_status'] = df.apply(
        lambda row: 'High_Volatility' if row['Volatility'] > row['Volatility_median'] else 'Low_Volatility', axis=1)
    return df


# Function to filter the input DataFrame based on the date range
def filter_date_range(df, start_date, end_date):
    date_mask = (df['date'] >= start_date) & (df['date'] <= end_date)
    return df[date_mask]


# Function to mark volatility trend in the input DataFrame
def mark_volatility_trend(df):
    df['Volatility_Trend'] = df.apply(
        lambda row: 'High Volatility + Trend' if row['Volatility_status'] == 'High_Volatility' and row[
            'trend'] == 1
        else ('High Volatility + No Trend' if row['Volatility_status'] == 'High_Volatility' and row['trend'] != 1
              else ('Low Volatility + Trend' if row['Volatility_status'] == 'Low_Volatility' and row['trend'] == 1
                    else 'Low Volatility + No Trend')), axis=1)
    return df


# Function to add spread columns to the input DataFrame
def add_spread_columns(df, spread_df, execution_df, symbol):
    spread_factors = {'EURUSD': 100000, 'GBPUSD': 100000, 'USDJPY': 1000, 'XAUUSD': 100}

    df = df.merge(spread_df.loc[spread_df['symbol'] == symbol, ['day', 'spread']], left_on='date', right_on='day',
                  how='left').drop(columns=['day'])
    df = df.rename(columns={'spread': 'typical_spread_in_points'})

    # Multiply the typical_spread column by the appropriate factor
    df['typical_spread_in_points'] = df['typical_spread_in_points'] * spread_factors[symbol]

    df = df.merge(execution_df.loc[execution_df['symbol_name'] == symbol, ['date', 'vol_avg_spread']], on='date',
                  how='left')
    df = df.rename(columns={'vol_avg_spread': 'weighted_avg_execution_spread_$'})

    return df


# Function to add PnL per lot column to the input DataFrame
def add_pnl_per_lot_column(df, pnl_df, symbol):
    df = df.merge(pnl_df.loc[pnl_df['symbol_name'] == symbol, ['day', 'PnL/Lot']], left_on='date', right_on='day',
                  how='left')
    df = df.rename(columns={'PnL/Lot': 'PnL_per_lot'})
    df.drop(columns=['day'], inplace=True)
    return df


# Modify the add_pnl_per_lot_column function to add the 'profit' and 'volume' columns
def add_pnl_per_lot_column(df, pnl_df, symbol):
    df = df.merge(pnl_df.loc[pnl_df['symbol_name'] == symbol, ['day', 'PnL/Lot', 'profit', 'volume']],
                  left_on='date', right_on='day', how='left')
    df = df.rename(columns={'PnL/Lot': 'PnL_per_lot', 'profit': 'total_profit', 'volume': 'total_volume'})
    df.drop(columns=['day'], inplace=True)
    return df


# Function to process input CSV files, add required columns, and save the processed data to a new CSV file
def process_csv(file_name, start_date, end_date, spread_df, execution_df, pnl_df_dict):
    symbol = file_name.split('/')[-1].split('_')[0]  # Extract the symbol from the file name
    df = pd.read_csv(file_name, parse_dates=['date'])
    df = add_volatility_columns(df)
    df = filter_date_range(df, start_date, end_date)
    df = mark_volatility_trend(df)
    df = add_spread_columns(df, spread_df, execution_df, symbol)
    df = add_pnl_per_lot_column(df, pnl_df_dict[symbol], symbol)
    output_dir = "processed_vol_trend_data"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    output_file_name = os.path.basename(file_name).replace("_chop", "_vol_trend")  # Replace '_chop' with '_vol_trend'
    output_file_path = os.path.join(output_dir, f'processed_{output_file_name}')
    df.to_csv(output_file_path, index=False)
    print(f"Processed {file_name}")


# Read input data from the spread and execution files
spread_df = pd.read_csv(spread_all_symbols_file, parse_dates=['day'])
execution_df = pd.read_csv(execution_spread_file, parse_dates=['date'])

# Process each input file
for file_name in file_names:
    process_csv(file_name, start_date, end_date, spread_df, execution_df, pnl_df_dict)


# Function to create pivot tables and charts using the processed data
def create_pivot_tables_and_charts(processed_files):
    processed_dfs = [pd.read_csv(file) for file in processed_files]

    combined_df = pd.concat(processed_dfs, keys=['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD'], names=['symbol']).reset_index(
        level=0).reset_index(drop=True)

    combined_df = combined_df[combined_df['Volatility_Trend'].notna()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivots"

    row_offset = 1

    for symbol in ['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD']:
        ws.cell(row=row_offset, column=1, value=symbol).font = Font(bold=True)

        symbol_df = combined_df[combined_df['symbol'] == symbol]

        pivot = pd.pivot_table(symbol_df, index='Volatility_Trend',
                               values=['typical_spread_in_points', 'weighted_avg_execution_spread_$',
                                       'PnL_per_lot', 'total_profit', 'total_volume'],
                               aggfunc={'typical_spread_in_points': 'mean',
                                        'weighted_avg_execution_spread_$': 'mean',
                                        'PnL_per_lot': ['mean', 'count'],
                                        'total_profit': 'sum',
                                        'total_volume': 'sum'})

        pivot.columns = ['_'.join(col).strip() if col[-1] == 'count' else col[0] for col in pivot.columns.values]
        pivot.rename(columns={'PnL_per_lot_count': 'count_of_occurrences'}, inplace=True)

        # Calculate the percentage of total occurrences for each Volatility_Trend
        total_occurrences = pivot['count_of_occurrences'].sum()
        pivot['percentage_of_occurrences'] = pivot['count_of_occurrences'] / total_occurrences * 100

        impact_values = []
        for trend in pivot.index:
            trend_df = symbol_df[symbol_df['Volatility_Trend'] == trend]

            # Drop rows with NaN values in the 'PnL_per_lot' column and 'weighted_avg_execution_spread_$' column
            trend_df = trend_df.dropna(subset=['PnL_per_lot', 'weighted_avg_execution_spread_$'])

            X = trend_df[['weighted_avg_execution_spread_$']]
            y = trend_df['PnL_per_lot']

            # Skip fitting the model if there are no valid data points
            if len(X) == 0 or len(y) == 0:
                impact_values.append(np.nan)
                continue

            model = LinearRegression()
            model.fit(X, y)

            impact = model.coef_[0]
            impact_values.append(impact)

        pivot['one_point_increase_of_weighted_spread_lr'] = impact_values

        row_offset += 1

        # Append the pivot table to the worksheet
        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws.append(r)

            row_offset += len(pivot) + 2

        # Adjust column widths to fit the content
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        wb.save("processed_vol_trend_data/pivots.xlsx")


# Add this line to the end of your script to call the function
create_pivot_tables_and_charts([
    "processed_vol_trend_data/processed_EURUSD_vol_trend.csv",
    "processed_vol_trend_data/processed_GBPUSD_vol_trend.csv",
    "processed_vol_trend_data/processed_USDJPY_vol_trend.csv",
    "processed_vol_trend_data/processed_XAUUSD_vol_trend.csv"
])
