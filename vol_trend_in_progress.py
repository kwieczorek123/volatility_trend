import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Define file names for input data
file_names = ['processed_CI_data/EURUSD_CI.csv', 'processed_CI_data/GBPUSD_CI.csv', 'processed_CI_data/USDJPY_CI.csv',
              'processed_CI_data/XAUUSD_CI.csv']
spread_all_symbols_file = 'vol_trend_data/spread_all_symbol.csv'
execution_spread_file = 'vol_trend_data/execution_spread.csv'
pnl_files = ['vol_trend_data/EURUSD_pnl.csv', 'vol_trend_data/GBPUSD_pnl.csv', 'vol_trend_data/USDJPY_pnl.csv',
             'vol_trend_data/XAUUSD_pnl.csv']

# Create a dictionary to store PnL data for each symbol
pnl_df_dict = {symbol: pd.read_csv(pnl_file, parse_dates=['day']) for pnl_file, symbol in
               zip(pnl_files, ['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD'])}

# Define date range and rolling window length
start_date = '2020-01-01'
end_date = '2023-03-30'
roll = 252
start_date_long = '2015-01-01'
end_date_long = '2022-12-31'


# Function to add volatility columns to the input DataFrame
def add_volatility_columns(df):
    df['Volatility'] = (df['high'] - df['low']) / df['low']
    df['Volatility_median'] = df['Volatility'].rolling(window=roll).median()
    df['Volatility_status'] = df.apply(
        lambda row: 'High_Volatility' if row['Volatility'] > row['Volatility_median'] else 'Low_Volatility', axis=1)
    return df


# Function to filter the input DataFrame based on the date range
def filter_date_range(df, start_date, end_date):
    date_mask = (df['date'] >= start_date) & (df['date'] <= end_date)
    return df[date_mask]


# Function to mark volatility trend in the input DataFrame
def mark_volatility_trend(df):
    df['Volatility_Trend'] = df.apply(
        lambda row: 'High Volatility + Trend' if row['Volatility_status'] == 'High_Volatility' and row['trend'] == 1
        else ('High Volatility + Trend False' if row['Volatility_status'] == 'High_Volatility' and row['trend'] == 0 and
                                                 row['trend_in_progress'] == 0.5
              else (
            'High Volatility + No Trend' if row['Volatility_status'] == 'High_Volatility' and row['trend'] == 0 and row[
                'trend_in_progress'] == 0
            else ('Low Volatility + Trend' if row['Volatility_status'] == 'Low_Volatility' and row['trend'] == 1
                  else (
                'Low Volatility + Trend False' if row['Volatility_status'] == 'Low_Volatility' and row['trend'] == 0 and
                                                  row['trend_in_progress'] == 0.5
                else 'Low Volatility + No Trend')))), axis=1)
    return df


# Function to add spread columns to the input DataFrame
def add_spread_columns(df, spread_df, execution_df, symbol):
    spread_factors = {'EURUSD': 100000, 'GBPUSD': 100000, 'USDJPY': 1000, 'XAUUSD': 100}

    df = df.merge(spread_df.loc[spread_df['symbol'] == symbol, ['day', 'spread']], left_on='date', right_on='day',
                  how='left').drop(columns=['day'])
    df = df.rename(columns={'spread': 'typical_spread_in_points'})

    # Multiply the typical_spread column by the appropriate factor
    df['typical_spread_in_points'] = df['typical_spread_in_points'] * spread_factors[symbol]

    df = df.merge(execution_df.loc[execution_df['symbol_name'] == symbol, ['date', 'vol_avg_spread']], on='date',
                  how='left')
    df = df.rename(columns={'vol_avg_spread': 'weighted_avg_execution_spread_$'})

    return df


# Function to add PnL per lot column to the input DataFrame
def add_pnl_per_lot_column(df, pnl_df, symbol):
    df = df.merge(pnl_df.loc[pnl_df['symbol_name'] == symbol, ['day', 'PnL/Lot', 'profit', 'volume']],
                  left_on='date', right_on='day', how='left')
    df = df.rename(columns={'PnL/Lot': 'PnL_per_lot', 'profit': 'total_profit', 'volume': 'total_volume'})
    df.drop(columns=['day'], inplace=True)
    return df


# Function to process input CSV files, add required columns, and save the processed data to a new CSV file
def process_csv(file_name, start_date, end_date, spread_df, execution_df, pnl_df_dict):
    symbol = file_name.split('/')[-1].split('_')[0]  # Extract the symbol from the file name
    df = pd.read_csv(file_name, parse_dates=['date'])
    df = add_volatility_columns(df)
    df = filter_date_range(df, start_date, end_date)
    df = mark_volatility_trend(df)
    df = add_spread_columns(df, spread_df, execution_df, symbol)
    df = add_pnl_per_lot_column(df, pnl_df_dict[symbol], symbol)
    output_dir = "processed_CI_test_data"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    output_file_name = os.path.basename(file_name).replace("_CI", "_vol_trend")  # Replace '_chop' with '_vol_trend'
    output_file_path = os.path.join(output_dir, f'processed_{output_file_name}')
    df.to_csv(output_file_path, index=False)
    print(f"Processed {file_name}")


def process_csv_long(file_name, start_date, end_date):
    symbol = file_name.split('/')[-1].split('_')[0]  # Extract the symbol from the file name
    df = pd.read_csv(file_name, parse_dates=['date'])
    df = add_volatility_columns(df)
    df = filter_date_range(df, start_date, end_date)
    df = mark_volatility_trend(df)

    output_dir = "processed_CI_test_data"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    output_file_name = os.path.basename(file_name).replace("_CI", "_vol_trend_long")
    output_file_path = os.path.join(output_dir, f'processed_{output_file_name}')
    df.to_csv(output_file_path, index=False)
    print(f"Processed {file_name}")


# Read input data from the spread and execution files
spread_df = pd.read_csv(spread_all_symbols_file, parse_dates=['day'])
execution_df = pd.read_csv(execution_spread_file, parse_dates=['date'])

# Process each input file
for file_name in file_names:
    process_csv(file_name, start_date, end_date, spread_df, execution_df, pnl_df_dict)

for file_name in file_names:
    process_csv_long(file_name, start_date_long, end_date_long)


# Function to create pivot tables and charts using the processed data
def create_pivot_tables_and_charts(processed_files):
    processed_dfs = [pd.read_csv(file) for file in processed_files]

    combined_df = pd.concat(processed_dfs, keys=['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD'], names=['symbol']).reset_index(
        level=0).reset_index(drop=True)

    combined_df = combined_df[combined_df['Volatility_Trend'].notna()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivots"

    row_offset = 1

    for symbol in ['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD']:
        ws.cell(row=row_offset, column=1, value=symbol).font = Font(bold=True)

        symbol_df = combined_df[combined_df['symbol'] == symbol]

        pivot = pd.pivot_table(symbol_df, index='Volatility_Trend',
                               values=['typical_spread_in_points', 'weighted_avg_execution_spread_$',
                                       'PnL_per_lot', 'total_profit', 'total_volume'],
                               aggfunc={'typical_spread_in_points': 'mean',
                                        'weighted_avg_execution_spread_$': 'mean',
                                        'PnL_per_lot': ['mean', 'count'],
                                        'total_profit': 'sum',
                                        'total_volume': 'sum'})

        pivot.columns = ['_'.join(col).strip() if col[-1] == 'count' else col[0] for col in pivot.columns.values]
        pivot.rename(columns={'PnL_per_lot_count': 'count_of_occurrences'}, inplace=True)

        # Calculate the percentage of total occurrences for each Volatility_Trend
        total_occurrences = pivot['count_of_occurrences'].sum()
        pivot['percentage_of_occurrences'] = pivot['count_of_occurrences'] / total_occurrences * 100

        # Change the calculation for the PnL_per_lot column
        pivot['PnL_per_lot'] = pivot['total_profit'] / pivot['total_volume']

        # Add the pct_impact_on_PnL_weighted_spread column
        pivot['pct_impact_on_PnL_exec_spread'] = abs(pivot['total_volume'] / pivot['total_profit'])

        # Rearrange the columns
        pivot = pivot.reindex(columns=['count_of_occurrences', 'percentage_of_occurrences', 'typical_spread_in_points',
                                       'weighted_avg_execution_spread_$', 'PnL_per_lot', 'total_profit',
                                       'total_volume', 'pct_impact_on_PnL_exec_spread'])

        # Calculate the percentage of total_profit and total_volume for each Volatility_Trend
        total_profit = pivot['total_profit'].sum()
        total_volume = pivot['total_volume'].sum()
        pivot['pct_total_profit'] = pivot['total_profit'] / total_profit * 100
        pivot['pct_total_volume'] = pivot['total_volume'] / total_volume * 100

        # Insert the new columns after total_profit and total_volume
        pivot = pivot.reindex(columns=['count_of_occurrences', 'percentage_of_occurrences', 'typical_spread_in_points',
                                       'weighted_avg_execution_spread_$', 'PnL_per_lot', 'total_profit',
                                       'pct_total_profit',
                                       'total_volume', 'pct_total_volume',
                                       'pct_impact_on_PnL_exec_spread'])

        row_offset += 1

        # Append the pivot table to the worksheet
        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws.append(r)

            row_offset += 1

        # Adjust column widths to fit the content
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        wb.save("processed_CI_test_data/pivots_CI_test.xlsx")


def create_pivot_tables_and_charts_long(processed_files):
    processed_dfs = [pd.read_csv(file) for file in processed_files]

    combined_df = pd.concat(processed_dfs, keys=['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD'], names=['symbol']).reset_index(
        level=0).reset_index(drop=True)

    combined_df = combined_df[combined_df['Volatility_Trend'].notna()]

    # Convert the 'date' column to a datetime object
    combined_df['date'] = pd.to_datetime(combined_df['date'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivots_long"

    row_offset = 1

    for symbol in ['EURUSD', 'GBPUSD', 'USDJPY', 'XAUUSD']:
        ws.cell(row=row_offset, column=1, value=symbol).font = Font(bold=True)

        symbol_df = combined_df[combined_df['symbol'] == symbol]

        # Get a list of years with no missing values in the Volatility_median column
        valid_years = symbol_df.groupby(symbol_df['date'].dt.year)['Volatility_median'].apply(lambda x: x.notna().all())
        valid_years = valid_years[valid_years].index.tolist()

        # Filter out the rows with invalid years
        symbol_df = symbol_df[symbol_df['date'].dt.year.isin(valid_years)]

        pivot = pd.pivot_table(symbol_df, index='Volatility_Trend', columns=symbol_df['date'].dt.year,
                               values='date', aggfunc='count')

        # Add a column for the total count of occurrences for all years
        pivot['total_count_of_occurrences'] = pivot.sum(axis=1)

        # Calculate percentage of total occurrences for each Volatility_Trend
        pivot['percentage_of_occurrences'] = (pivot['total_count_of_occurrences'] / pivot[
            'total_count_of_occurrences'].sum()) * 100

        row_offset += 1

        # Append the pivot table to the worksheet
        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws.append(r)

            row_offset += 1

        # Adjust column widths to fit the content
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save("processed_CI_test_data/pivots_long_CI_test.xlsx")


# Add this line to the end of your script to call the function
create_pivot_tables_and_charts([
    "processed_CI_test_data/processed_EURUSD_vol_trend.csv",
    "processed_CI_test_data/processed_GBPUSD_vol_trend.csv",
    "processed_CI_test_data/processed_USDJPY_vol_trend.csv",
    "processed_CI_test_data/processed_XAUUSD_vol_trend.csv"
])

create_pivot_tables_and_charts_long([
    "processed_CI_test_data/processed_EURUSD_vol_trend_long.csv",
    "processed_CI_test_data/processed_GBPUSD_vol_trend_long.csv",
    "processed_CI_test_data/processed_USDJPY_vol_trend_long.csv",
    "processed_CI_test_data/processed_XAUUSD_vol_trend_long.csv"
])
